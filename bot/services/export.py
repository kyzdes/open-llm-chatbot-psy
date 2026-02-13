"""Export helpers: Markdown → PDF (fpdf2) and Markdown tables → Excel (openpyxl)."""

from __future__ import annotations

import io
import logging
import os
import re

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# PDF generation (fpdf2 + DejaVu for Cyrillic)
# ---------------------------------------------------------------------------

_DEJAVU_DIR = os.path.join(os.path.dirname(__file__), "fonts")
_DEJAVU_REGULAR = os.path.join(_DEJAVU_DIR, "DejaVuSans.ttf")
_DEJAVU_BOLD = os.path.join(_DEJAVU_DIR, "DejaVuSans-Bold.ttf")


def _has_dejavu() -> bool:
    return os.path.isfile(_DEJAVU_REGULAR)


def markdown_to_pdf(text: str, title: str = "Document") -> io.BytesIO | None:
    """Convert markdown text to a PDF BytesIO.

    Returns None if generation fails (caller should fall back to .md).
    """
    try:
        from fpdf import FPDF  # fpdf2
    except ImportError:
        logger.warning("fpdf2 not installed, PDF export unavailable")
        return None

    try:
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        # Register DejaVu font for Cyrillic support
        if _has_dejavu():
            pdf.add_font("DejaVu", "", _DEJAVU_REGULAR)
            pdf.add_font("DejaVu", "B", _DEJAVU_BOLD)
            font_family = "DejaVu"
        else:
            # Fallback — built-in Helvetica (no Cyrillic, but at least won't crash)
            font_family = "Helvetica"
            logger.warning("DejaVu fonts not found at %s, Cyrillic may not render", _DEJAVU_DIR)

        pdf.set_font(font_family, "B", 16)
        pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(4)

        _mc = dict(new_x="LMARGIN", new_y="NEXT")  # reset cursor after each multi_cell

        for line in text.split("\n"):
            stripped = line.strip()

            # Headings
            if stripped.startswith("## "):
                pdf.ln(4)
                pdf.set_font(font_family, "B", 13)
                pdf.multi_cell(0, 7, stripped[3:], **_mc)
                pdf.ln(2)
                continue
            if stripped.startswith("# "):
                pdf.ln(4)
                pdf.set_font(font_family, "B", 15)
                pdf.multi_cell(0, 8, stripped[2:], **_mc)
                pdf.ln(2)
                continue
            if stripped.startswith("### "):
                pdf.ln(2)
                pdf.set_font(font_family, "B", 11)
                pdf.multi_cell(0, 6, stripped[4:], **_mc)
                pdf.ln(1)
                continue

            # Bold lines (e.g. **text**)
            cleaned = re.sub(r"\*\*(.+?)\*\*", r"\1", stripped)
            is_bold = cleaned != stripped

            # Bullet points
            if stripped.startswith("- ") or stripped.startswith("* "):
                pdf.set_font(font_family, "B" if is_bold else "", 10)
                pdf.multi_cell(0, 6, "  \u2022 " + cleaned[2:], **_mc)
                continue

            # Empty line
            if not stripped:
                pdf.ln(3)
                continue

            # Regular paragraph
            pdf.set_font(font_family, "B" if is_bold else "", 10)
            pdf.multi_cell(0, 6, cleaned, **_mc)

        buf = io.BytesIO()
        pdf.output(buf)
        buf.seek(0)
        return buf

    except Exception:
        logger.exception("PDF generation failed")
        return None


# ---------------------------------------------------------------------------
# Excel generation (openpyxl) — extracts markdown tables
# ---------------------------------------------------------------------------

_TABLE_ROW_RE = re.compile(r"^\|(.+)\|$")
_SEPARATOR_RE = re.compile(r"^[\|\s:\-]+$")


def _extract_markdown_tables(text: str) -> list[list[list[str]]]:
    """Parse all markdown tables from text.

    Returns list of tables, each table is list of rows, each row is list of cell strings.
    """
    tables: list[list[list[str]]] = []
    current_table: list[list[str]] = []

    for line in text.split("\n"):
        line = line.strip()
        m = _TABLE_ROW_RE.match(line)
        if m:
            # Check if this is a separator row (|---|---|...)
            if _SEPARATOR_RE.match(line):
                continue
            cells = [c.strip() for c in m.group(1).split("|")]
            current_table.append(cells)
        else:
            if current_table:
                tables.append(current_table)
                current_table = []

    if current_table:
        tables.append(current_table)

    return tables


def markdown_tables_to_excel(text: str, title: str = "Report") -> io.BytesIO | None:
    """Extract markdown tables from text and convert to .xlsx.

    Returns None if no tables found or generation fails.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        logger.warning("openpyxl not installed, Excel export unavailable")
        return None

    tables = _extract_markdown_tables(text)
    if not tables:
        logger.info("No markdown tables found for Excel export")
        return None

    try:
        wb = Workbook()

        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        cell_alignment = Alignment(wrap_text=True, vertical="top")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for idx, table in enumerate(tables):
            if idx == 0:
                ws = wb.active
                ws.title = title[:31]  # Excel sheet name limit
            else:
                ws = wb.create_sheet(title=f"Table {idx + 1}")

            for row_idx, row in enumerate(table):
                for col_idx, cell_value in enumerate(row):
                    cell = ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
                    cell.alignment = cell_alignment
                    cell.border = thin_border

                    if row_idx == 0:
                        cell.font = header_font_white
                        cell.fill = header_fill
                    else:
                        # Try to convert numeric values
                        try:
                            num = float(cell_value.replace(",", ".").replace(" ", "").replace("\u00a0", "").rstrip("%"))
                            if cell_value.strip().endswith("%"):
                                cell.value = num / 100
                                cell.number_format = "0.0%"
                            else:
                                cell.value = num
                                if num == int(num):
                                    cell.number_format = "#,##0"
                                else:
                                    cell.number_format = "#,##0.00"
                        except (ValueError, AttributeError):
                            pass

            # Auto-width columns
            for col_idx in range(1, (len(table[0]) if table else 0) + 1):
                max_len = 0
                for row in table:
                    if col_idx - 1 < len(row):
                        max_len = max(max_len, len(row[col_idx - 1]))
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max(max_len + 2, 10), 50)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    except Exception:
        logger.exception("Excel generation failed")
        return None


def export_as_markdown(text: str, title: str = "document") -> io.BytesIO:
    """Fallback: wrap text into a .md BytesIO."""
    buf = io.BytesIO()
    buf.write(text.encode("utf-8"))
    buf.seek(0)
    return buf
